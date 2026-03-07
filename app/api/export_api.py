from fastapi import APIRouter, Request, Depends, HTTPException, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, select, desc
from typing import Optional, List
import json
import zipfile
import io
import os
from datetime import datetime

# Internal Imports
from app.core.database import get_session
from app.models import Institution, SystemSnapshot, User
from app.logic.auth import get_current_user
from app.logic.permissions import get_institution_with_access
from app.utils.exporting import (
    export_all_institutions_bundle,
    export_institution_to_csv_zip,
    export_institution_to_json,
    export_institutions_bundle,
    export_institution_to_excel,
    import_institution_from_json,
)

router = APIRouter()
from fastapi import Response
from app.utils.context import TemplateResponse
from app.logic.auth import get_current_user

# --- 1. ZIP Restore Logic ---
def _handle_zip_restore(session: Session, file_obj, specific_institution: Optional[Institution] = None):
    """ZIP فائل سے ڈیٹا کی بحالی کا منطقی فنکشن۔"""
    restored_count = 0
    errors = []
    try:
        wrap_file = io.BytesIO(file_obj) if isinstance(file_obj, bytes) else file_obj
        with zipfile.ZipFile(wrap_file) as z:
            json_files = [f for f in z.namelist() if f.endswith(".json")]
            for json_path in json_files:
                slug_candidate = json_path.split('/')[0]
                if specific_institution and slug_candidate != specific_institution.slug:
                    continue
                institution = session.exec(select(Institution).where(Institution.slug == slug_candidate)).first()
                if institution:
                    json_content = z.read(json_path).decode('utf-8')
                    result = import_institution_from_json(institution, session, json_content)
                    if "error" in result: errors.append(f"{slug_candidate}: {result['error']}")
                    else: restored_count += 1
        return restored_count, errors
    except Exception as e:
        return 0, [str(e)]

# --- 2. Backup Manager ---
@router.get("/backups/", response_class=HTMLResponse, name="backup_manager")
@router.get("/{institution_slug}/backups/", response_class=HTMLResponse, name="institution_backup_manager")
async def backup_manager(request: Request, institution_slug: Optional[str] = None, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if current_user.is_superuser and not institution_slug:
        snapshots = session.exec(select(SystemSnapshot).order_by(desc(SystemSnapshot.created_at))).all()
        current_inst = None
    else:
        current_inst, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
        snapshots = session.exec(select(SystemSnapshot).where(SystemSnapshot.inst_id == current_inst.id).order_by(desc(SystemSnapshot.created_at))).all()
    
    return await TemplateResponse.render("dms/backup_manager.html", request, session, {
        "snapshots": snapshots, 
        "institution": current_inst
    })

# --- 3. Snapshot Actions ---
@router.get("/backups/create/", name="create_manual_snapshot")
@router.get("/{institution_slug}/backups/create/", name="create_institution_snapshot")
async def create_manual_snapshot(request: Request, institution_slug: Optional[str] = None, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    target_slug = institution_slug or request.query_params.get('inst_slug')
    # Logic to create snapshot should be here or in logic layer
    return RedirectResponse(url=f"/{target_slug}/backups" if target_slug else "/backups", status_code=303)

@router.post("/backups/restore/{snapshot_id}/", name="restore_snapshot")
async def restore_snapshot(snapshot_id: int, institution_slug: Optional[str] = None, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if not current_user.is_superuser: raise HTTPException(status_code=403)
    # Restore logic
    return RedirectResponse(url=f"/{institution_slug}/backups" if institution_slug else "/backups", status_code=303)

@router.post("/backups/delete/{snapshot_id}/", name="delete_snapshot")
async def delete_snapshot_route(snapshot_id: int, institution_slug: Optional[str] = None, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if not current_user.is_superuser: raise HTTPException(status_code=403)
    snapshot = session.get(SystemSnapshot, snapshot_id)
    if snapshot:
        session.delete(snapshot)
        session.commit()
    return RedirectResponse(url=f"/{institution_slug}/backups" if institution_slug else "/backups", status_code=303)

@router.get("/backups/download/{snapshot_id}/", name="download_snapshot")
async def download_snapshot_file(snapshot_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    snapshot = session.get(SystemSnapshot, snapshot_id)
    if not snapshot: raise HTTPException(status_code=404)
    if not current_user.is_superuser and snapshot.inst_id:
        inst = session.get(Institution, snapshot.inst_id)
        if inst.user_id != current_user.id: raise HTTPException(status_code=403)
    return FileResponse(snapshot.file_path, filename=os.path.basename(snapshot.file_path))

# --- 3.1 Django-Compatible Aliases ---
@router.post("/{institution_slug}/backups/restore/{snapshot_id}/", name="restore_snapshot_ajax")
async def restore_snapshot_ajax(institution_slug: str, snapshot_id: int,
                                session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if not current_user.is_superuser: raise HTTPException(status_code=403)
    return RedirectResponse(url=f"/{institution_slug}/backups/", status_code=303)

@router.get("/{institution_slug}/backups/download/{snapshot_id}/", name="download_snapshot_file")
async def download_snapshot_file_scoped(snapshot_id: int, institution_slug: str,
                                       session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    snapshot = session.get(SystemSnapshot, snapshot_id)
    if not snapshot: raise HTTPException(status_code=404)
    return FileResponse(snapshot.file_path, filename=os.path.basename(snapshot.file_path))

@router.get("/{institution_slug}/export/", name="institution_export")
async def institution_export_route(institution_slug: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    institution, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
    json_payload = export_institution_to_json(institution, session)
    from fastapi.responses import Response
    return Response(content=json_payload, media_type="application/json",
                    headers={"Content-Disposition": f'attachment; filename="{institution.slug}.json"'})

# --- 4. Exports ---
@router.get("/{institution_slug}/export/json/", name="export_institution_json")
async def download_institution_export_json_route(institution_slug: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    institution, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
    json_payload = export_institution_to_json(institution, session)
    return Response(content=json_payload, media_type="application/json", headers={"Content-Disposition": f'attachment; filename="{institution.slug}.json"'})

@router.get("/{institution_slug}/export/bundle/", name="export_institution_bundle")
async def download_institution_export_bundle_route(request: Request, institution_slug: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    institution, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
    if request.query_params.get("format") == "sheet":
        excel_bytes = export_institution_to_excel(institution, session)
        return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{institution.slug}.xlsx"'})
    archive = export_institutions_bundle([institution], session)
    return Response(content=archive, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{institution.slug}-bundle.zip"'})

@router.get("/{institution_slug}/export/sheet/", name="export_institution_sheet")
async def download_institution_export_sheet_route(institution_slug: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    institution, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
    excel_bytes = export_institution_to_excel(institution, session)
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{institution.slug}.xlsx"'})

# --- 5. Restores ---
@router.post("/{institution_slug}/restore/", name="restore_institution")
async def restore_institution_backup(institution_slug: str, backup_file: UploadFile = File(...), session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    institution, access = get_institution_with_access(institution_slug, session, current_user, access_type='admin')
    content = await backup_file.read()
    if backup_file.filename.endswith(".json"):
        import_institution_from_json(institution, session, content.decode('utf-8'))
    elif backup_file.filename.endswith(".zip"):
        _handle_zip_restore(session, content, specific_institution=institution)
    return RedirectResponse(url=f"/{institution_slug}/settings", status_code=303)

@router.post("/system/restore/", name="restore_system")
async def restore_system_backup(backup_file: UploadFile = File(...), session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if not current_user.is_superuser: raise HTTPException(status_code=403)
    content = await backup_file.read()
    if backup_file.filename.endswith(".zip"):
        _handle_zip_restore(session, content)
    return RedirectResponse(url="/backups", status_code=303)

@router.get("/system/export/all/", name="export_all_institutions")
async def download_all_institutions_export_route(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    if not current_user.is_superuser: raise HTTPException(status_code=403)
    archive = export_all_institutions_bundle(session)
    return Response(content=archive, media_type="application/zip", headers={"Content-Disposition": 'attachment; filename="system-full-backup.zip"'})


# ─────────────────────────────────────────────────────────────────────────────
# 6. Attendance Export — Excel & CSV
# ─────────────────────────────────────────────────────────────────────────────

def _build_attendance_excel(institution, session_db, course_id: Optional[int] = None,
                             month: Optional[int] = None, year: Optional[int] = None) -> bytes:
    """
    ماہانہ حاضری رپورٹ Excel میں بنانا۔
    rows = طلبہ،  columns = دنوں کی تاریخیں
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as dt
    from sqlmodel import select
    from app.models import Student, Admission, Attendance, ClassSession, Course, DailyAttendance
    import calendar

    today = dt.today()
    yr = year or today.year
    mn = month or today.month
    start_date = dt(yr, mn, 1)
    _, last_day = calendar.monthrange(yr, mn)
    end_date = dt(yr, mn, last_day)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[mn][:3]} {yr}"
    ws.sheet_view.rightToLeft = True

    hdr_fill   = PatternFill("solid", fgColor="1E293B")
    present_fill = PatternFill("solid", fgColor="166534")
    absent_fill  = PatternFill("solid", fgColor="991B1B")
    late_fill    = PatternFill("solid", fgColor="92400E")
    hdr_font   = Font(bold=True, color="F1F5F9", size=9)
    thin       = Side(style='thin', color="334155")
    brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

    if course_id:
        courses = [c for c in [session_db.get(Course, course_id)] if c]
    else:
        courses = session_db.exec(
            select(Course).where(Course.inst_id == institution.id, Course.is_active == True)
        ).all()

    row_idx = 1
    for course in courses:
        # Course heading
        cell = ws.cell(row=row_idx, column=1, value=f"کورس: {course.title}  |  {calendar.month_name[mn]} {yr}")
        cell.font = Font(bold=True, size=11, color="60A5FA")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_day + 4)
        row_idx += 1

        students_q = session_db.exec(
            select(Student).join(Admission).where(
                Admission.course_id == course.id,
                Student.inst_id == institution.id,
                Student.deleted_at == None
            ).distinct()
        ).all()

        if not students_q:
            ws.cell(row=row_idx, column=1, value="(کوئی طالب علم نہیں)").font = Font(italic=True, color="64748B")
            row_idx += 2
            continue

        # Header
        headers = ["نام", "رول نمبر"] + [str(d) for d in range(1, last_day + 1)] + ["حاضر", "%"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center'); c.border = brd
        row_idx += 1

        for student in students_q:
            att_records = session_db.exec(
                select(Attendance, ClassSession).join(
                    ClassSession, Attendance.session_id == ClassSession.id
                ).where(
                    Attendance.student_id == student.id,
                    ClassSession.course_id == course.id,
                    ClassSession.date >= start_date, ClassSession.date <= end_date
                )
            ).all()
            day_map = {cs.date.day: att.status for att, cs in att_records}

            daily_records = session_db.exec(
                select(DailyAttendance).where(
                    DailyAttendance.student_id == student.id,
                    DailyAttendance.date >= start_date, DailyAttendance.date <= end_date
                )
            ).all()
            for da in daily_records:
                if da.date.day not in day_map: day_map[da.date.day] = da.status

            present_count = sum(1 for s in day_map.values() if s == 'present')
            total_days    = len(day_map)
            pct = f"{round(present_count/total_days*100, 1)}%" if total_days else "—"

            row_vals = [student.full_name, student.reg_id or "—"]
            row_vals += [day_map.get(d) for d in range(1, last_day + 1)]
            row_vals += [present_count, pct]

            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.alignment = Alignment(horizontal='center'); c.border = brd
                if col > 2 and col <= last_day + 2 and val:
                    if   val == 'present': c.fill = present_fill; c.font = Font(color="FFFFFF", bold=True); c.value = "✓"
                    elif val == 'absent':  c.fill = absent_fill;  c.font = Font(color="FFFFFF", bold=True); c.value = "✗"
                    elif val == 'late':    c.fill = late_fill;    c.font = Font(color="FFFFFF", bold=True); c.value = "L"
                    else: c.value = ""
            row_idx += 1
        row_idx += 2

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    for col in range(3, last_day + 5):
        ws.column_dimensions[get_column_letter(col)].width = 4 if col < last_day + 3 else 8

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_attendance_csv(institution, session_db, course_id=None, month=None, year=None) -> str:
    import csv as _csv, io as _io
    from datetime import date as dt
    from sqlmodel import select
    from app.models import Student, Admission, Attendance, ClassSession, DailyAttendance, Course
    import calendar

    today = dt.today()
    yr = year or today.year; mn = month or today.month
    start_date = dt(yr, mn, 1)
    _, last_day = calendar.monthrange(yr, mn)
    end_date = dt(yr, mn, last_day)

    buf = _io.StringIO()
    w = _csv.writer(buf)

    if course_id:
        courses = [c for c in [session_db.get(Course, course_id)] if c]
    else:
        courses = session_db.exec(select(Course).where(Course.inst_id == institution.id, Course.is_active == True)).all()

    for course in courses:
        w.writerow([f"Course: {course.title}", f"Month: {mn}/{yr}"])
        w.writerow(["Name", "Reg ID"] + list(range(1, last_day+1)) + ["Present", "Percent"])
        students_q = session_db.exec(
            select(Student).join(Admission).where(
                Admission.course_id == course.id, Student.inst_id == institution.id, Student.deleted_at == None
            ).distinct()
        ).all()
        for s in students_q:
            atts = session_db.exec(
                select(Attendance, ClassSession).join(ClassSession, Attendance.session_id == ClassSession.id).where(
                    Attendance.student_id == s.id, ClassSession.course_id == course.id,
                    ClassSession.date >= start_date, ClassSession.date <= end_date)
            ).all()
            dm = {cs.date.day: att.status for att, cs in atts}
            for da in session_db.exec(select(DailyAttendance).where(DailyAttendance.student_id == s.id, DailyAttendance.date >= start_date, DailyAttendance.date <= end_date)).all():
                if da.date.day not in dm: dm[da.date.day] = da.status
            pc = sum(1 for v in dm.values() if v == 'present')
            pct = f"{round(pc/len(dm)*100,1)}%" if dm else "0%"
            w.writerow([s.full_name, s.reg_id or ""] + [dm.get(d, "") for d in range(1, last_day+1)] + [pc, pct])
        w.writerow([])
    return buf.getvalue()


@router.get("/{institution_slug}/export/attendance/", name="export_attendance")
async def export_attendance(
    institution_slug: str,
    fmt: str = "xlsx",
    course_id: Optional[int] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """
    ماہانہ حاضری رپورٹ Download کریں۔
    ?fmt=xlsx  (default)  یا  ?fmt=csv
    ?course_id=5  (اختیاری)
    ?month=3&year=2026
    """
    from fastapi import Response as R
    import calendar
    from datetime import date
    institution, _ = get_institution_with_access(institution_slug, session, current_user, access_type='academic_view')

    today = date.today()
    mn = month or today.month; yr = year or today.year
    sfx = f"_c{course_id}" if course_id else "_all"
    fname = f"{institution.slug}_attendance_{yr}_{mn:02d}{sfx}"

    if fmt == "csv":
        data = _build_attendance_csv(institution, session, course_id=course_id, month=mn, year=yr)
        return R(content=data.encode('utf-8-sig'), media_type="text/csv",
                 headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'})

    data = _build_attendance_excel(institution, session, course_id=course_id, month=mn, year=yr)
    return R(content=data,
             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
             headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


